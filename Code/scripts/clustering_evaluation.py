import warnings
import pandas as pd
import numpy as np
import json
import logging
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any, Union
from dataclasses import dataclass
import pickle
from datetime import datetime
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.chart import BarChart, ScatterChart, LineChart, Reference
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import plotly.offline as pyo

# Scikit-learn imports
from sklearn.cluster import KMeans, DBSCAN, AgglomerativeClustering
from sklearn.decomposition import PCA, LatentDirichletAllocation
from sklearn.metrics import silhouette_score, calinski_harabasz_score, davies_bouldin_score
from sklearn.preprocessing import StandardScaler

# UMAP and Fuzzy C-Means
try:
    import umap.umap_ as umap
except ImportError:
    umap = None

try:
    from skfuzzy import cluster as fuzz
    import skfuzzy as fuzz_module
except ImportError:
    fuzz = None
    fuzz_module = None

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


@dataclass
class ClusteringResult:
    """Data class to store clustering results."""
    algorithm: str
    reduction_method: str
    n_components: Optional[int]
    k_value: Optional[int]
    cluster_labels: np.ndarray
    reduced_embeddings: Optional[np.ndarray]
    original_embeddings: np.ndarray
    silhouette_score: float
    calinski_harabasz_score: float
    davies_bouldin_score: float
    n_clusters_found: int
    execution_time: float
    parameters: Dict[str, Any]
    combination_id: str  # New field for unique identification


def generate_combination_id(algorithm: str, reduction_method: str,
                            n_components: Optional[int], k_value: Optional[int],
                            algorithm_params: Optional[Dict] = None) -> str:
    """Generate unique ID for each combination."""
    components_str = f"comp{n_components}" if n_components else "compNone"
    k_str = f"k{k_value}" if k_value else "kNone"

    if algorithm_params and algorithm == 'dbscan':
        params_str = f"eps{algorithm_params.get('eps', 0.5)}_min{algorithm_params.get('min_samples', 5)}"
        return f"{algorithm}_{reduction_method}_{components_str}_{params_str}"

    return f"{algorithm}_{reduction_method}_{components_str}_{k_str}"


def load_embeddings_data(csv_path: str, embeddings_column: str = 'embeddings',
                         text_column: str = 'text') -> Tuple[pd.DataFrame, np.ndarray]:
    """
    Load dataset with embeddings from CSV file.

    Args:
        csv_path: Path to CSV file containing embeddings
        embeddings_column: Name of column containing embeddings
        text_column: Name of column containing text data

    Returns:
        Tuple[pd.DataFrame, np.ndarray]: DataFrame and embeddings array
    """
    df = pd.read_csv(csv_path)

    if embeddings_column not in df.columns:
        raise ValueError(f"Column '{embeddings_column}' not found in dataset")

    if text_column not in df.columns:
        logger.warning(f"Text column '{text_column}' not found. Will use index for text output.")
        df[text_column] = df.index.astype(str)

    # Convert JSON embeddings to numpy array
    embeddings = np.array([json.loads(emb) for emb in df[embeddings_column]])

    logger.info(f"Loaded dataset: {len(df)} samples, embedding dimension: {embeddings.shape[1]}")
    return df, embeddings


def apply_dimensionality_reduction(embeddings: np.ndarray,
                                   method: str,
                                   n_components: Optional[int] = None,
                                   random_state: int = 42) -> Tuple[np.ndarray, Dict[str, Any]]:
    """Apply dimensionality reduction to embeddings."""
    if method.lower() == 'none':
        return embeddings, {'method': 'none'}

    start_time = datetime.now()

    if method.lower() == 'pca':
        if n_components is None:
            n_components = min(50, embeddings.shape[1])

        reducer = PCA(n_components=n_components, random_state=random_state)
        reduced_embeddings = reducer.fit_transform(embeddings)

        params = {
            'method': 'pca',
            'n_components': n_components,
            'explained_variance_ratio': reducer.explained_variance_ratio_.sum(),
            'random_state': random_state
        }

    elif method.lower() == 'umap':
        if umap is None:
            raise ImportError("UMAP not installed. Install with: pip install umap-learn")

        if n_components is None:
            n_components = 2

        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", message="n_jobs value .* overridden")

            reducer = umap.UMAP(
                n_components=n_components,
                random_state=random_state,
                n_neighbors=15,
                min_dist=0.1
            )
            reduced_embeddings = reducer.fit_transform(embeddings)

        params = {
            'method': 'umap',
            'n_components': n_components,
            'n_neighbors': 15,
            'min_dist': 0.1,
            'random_state': random_state
        }

    else:
        raise ValueError(f"Unknown reduction method: {method}")

    execution_time = (datetime.now() - start_time).total_seconds()
    params['execution_time'] = execution_time

    logger.info(f"Applied {method.upper()} reduction: {embeddings.shape} -> {reduced_embeddings.shape}")
    return reduced_embeddings, params


def apply_kmeans_clustering(embeddings: np.ndarray, k: int, random_state: int = 42) -> Tuple[
    np.ndarray, Dict[str, Any]]:
    """Apply K-means clustering."""
    start_time = datetime.now()

    kmeans = KMeans(n_clusters=k, random_state=random_state, n_init=10)
    labels = kmeans.fit_predict(embeddings)

    execution_time = (datetime.now() - start_time).total_seconds()

    params = {
        'algorithm': 'kmeans',
        'n_clusters': k,
        'random_state': random_state,
        'inertia': kmeans.inertia_,
        'n_iter': kmeans.n_iter_,
        'execution_time': execution_time
    }

    return labels, params


def apply_fuzzy_cmeans_clustering(embeddings: np.ndarray, k: int, random_state: int = 42,
                                  membership_threshold: float = 0.3) -> Tuple[np.ndarray, Dict[str, Any]]:
    """Apply Fuzzy C-means clustering with multiple membership support."""
    if fuzz is None:
        raise ImportError("scikit-fuzzy not installed. Install with: pip install scikit-fuzzy")

    start_time = datetime.now()

    # Transpose data for scikit-fuzzy format
    data_T = embeddings.T

    # Apply fuzzy c-means
    cntr, u, u0, d, jm, p, fpc = fuzz.cmeans(
        data_T, k, 2, error=0.005, maxiter=1000, init=None
    )

    # Handle multiple membership based on threshold
    labels = []
    multiple_membership_count = 0

    for i in range(u.shape[1]):  # For each data point
        memberships = u[:, i]
        clusters_above_threshold = np.where(memberships >= membership_threshold)[0]

        if len(clusters_above_threshold) == 0:
            # No cluster meets threshold, assign to highest membership
            labels.append([np.argmax(memberships)])
        else:
            # Assign to all clusters above threshold
            labels.append(clusters_above_threshold.tolist())
            if len(clusters_above_threshold) > 1:
                multiple_membership_count += 1

    execution_time = (datetime.now() - start_time).total_seconds()

    params = {
        'algorithm': 'fuzzy_cmeans',
        'n_clusters': k,
        'membership_threshold': membership_threshold,
        'fpc': fpc,  # Fuzzy partition coefficient
        'jm': jm[-1],  # Final objective function value
        'n_iterations': len(jm),
        'multiple_membership_count': multiple_membership_count,
        'execution_time': execution_time
    }

    return labels, params


def apply_dbscan_clustering(embeddings: np.ndarray, eps: float = 0.5, min_samples: int = 5) -> Tuple[
    np.ndarray, Dict[str, Any]]:
    """Apply DBSCAN clustering."""
    start_time = datetime.now()

    dbscan = DBSCAN(eps=eps, min_samples=min_samples)
    labels = dbscan.fit_predict(embeddings)

    execution_time = (datetime.now() - start_time).total_seconds()

    n_clusters = len(set(labels)) - (1 if -1 in labels else 0)
    n_noise = list(labels).count(-1)

    params = {
        'algorithm': 'dbscan',
        'eps': eps,
        'min_samples': min_samples,
        'n_clusters_found': n_clusters,
        'n_noise_points': n_noise,
        'execution_time': execution_time
    }

    return labels, params


def apply_lda_clustering(embeddings: np.ndarray, k: int, random_state: int = 42) -> Tuple[np.ndarray, Dict[str, Any]]:
    """Apply LDA topic modeling as clustering."""
    start_time = datetime.now()

    # Convert embeddings to non-negative values for LDA
    embeddings_min = embeddings.min()
    if embeddings_min < 0:
        embeddings_shifted = embeddings - embeddings_min + 1e-10
    else:
        embeddings_shifted = embeddings + 1e-10
    embeddings_scaled = embeddings_shifted * 1000
    embeddings_lda = np.maximum(embeddings_scaled.astype(np.int32), 1)

    lda = LatentDirichletAllocation(
        n_components=k,
        random_state=random_state,
        max_iter=50,
        learning_method='online',
        learning_offset=10,
        learning_decay=0.7,
        batch_size=128,
        n_jobs=-1 if embeddings.shape[0] > 500 else 1
    )

    doc_topic_dist = lda.fit_transform(embeddings_lda)
    labels = np.argmax(doc_topic_dist, axis=1)

    execution_time = (datetime.now() - start_time).total_seconds()

    params = {
        'algorithm': 'lda',
        'n_components': k,
        'random_state': random_state,
        'perplexity': lda.perplexity(embeddings_lda),
        'log_likelihood': lda.score(embeddings_lda),
        'n_iter': lda.n_iter_,
        'execution_time': execution_time
    }

    return labels, params


def apply_agglomerative_clustering(embeddings: np.ndarray, k: int, random_state: int = 42) -> Tuple[np.ndarray, Dict[str, Any]]:
    """Apply Agglomerative Hierarchical clustering."""
    start_time = datetime.now()

    agglomerative = AgglomerativeClustering(n_clusters=k, linkage='ward', metric='euclidean')

    labels = agglomerative.fit_predict(embeddings)
    execution_time = (datetime.now() - start_time).total_seconds()

    params = {
        'algorithm': 'agglomerative',
        'n_clusters': k,
        'linkage': 'ward',
        'metric': 'euclidean',
        'execution_time': execution_time
    }
    return labels, params


def calculate_clustering_metrics(embeddings: np.ndarray, labels: Union[np.ndarray, List]) -> Tuple[float, float, float]:
    """Calculate intrinsic clustering evaluation metrics."""

    # Handle multiple membership case (convert to single membership for metrics)
    if isinstance(labels, list) and isinstance(labels[0], list):
        # For multiple membership, use the first assigned cluster for metrics calculation
        single_labels = np.array([label_list[0] for label_list in labels])
    else:
        single_labels = labels

    # Remove noise points for DBSCAN (label -1)
    mask = single_labels != -1
    if mask.sum() < 2:  # Need at least 2 points
        return -1.0, 0.0, float('inf')

    clean_embeddings = embeddings[mask]
    clean_labels = single_labels[mask]

    # Need at least 2 clusters for meaningful metrics
    n_clusters = len(set(clean_labels))
    if n_clusters < 2:
        return -1.0, 0.0, float('inf')

    try:
        silhouette = silhouette_score(clean_embeddings, clean_labels)
        calinski_harabasz = calinski_harabasz_score(clean_embeddings, clean_labels)
        davies_bouldin = davies_bouldin_score(clean_embeddings, clean_labels)

        return silhouette, calinski_harabasz, davies_bouldin
    except Exception as e:
        logger.warning(f"Error calculating metrics: {str(e)}")
        return -1.0, 0.0, float('inf')


def run_clustering_combination(embeddings: np.ndarray,
                               algorithm: str,
                               reduction_method: str,
                               n_components: Optional[int] = None,
                               k_value: Optional[int] = None,
                               algorithm_params: Optional[Dict] = None) -> ClusteringResult:
    """Run a single clustering combination."""
    start_time = datetime.now()

    if algorithm_params is None:
        algorithm_params = {}

    # Generate combination ID
    combination_id = generate_combination_id(algorithm, reduction_method, n_components, k_value, algorithm_params)

    # Apply dimensionality reduction
    if reduction_method.lower() != 'none':
        reduced_embeddings, reduction_params = apply_dimensionality_reduction(
            embeddings, reduction_method, n_components
        )
    else:
        reduced_embeddings = embeddings
        reduction_params = {'method': 'none'}

    # Apply clustering algorithm
    if algorithm.lower() == 'kmeans':
        if k_value is None:
            raise ValueError("k_value required for K-means")
        labels, clustering_params = apply_kmeans_clustering(reduced_embeddings, k_value)

    elif algorithm.lower() == 'fuzzy_cmeans':
        if k_value is None:
            raise ValueError("k_value required for Fuzzy C-means")
        # Extract membership threshold from algorithm_params if provided
        membership_threshold = algorithm_params.get('membership_threshold', 0.3) if algorithm_params else 0.3
        labels, clustering_params = apply_fuzzy_cmeans_clustering(reduced_embeddings, k_value, membership_threshold=membership_threshold)

    elif algorithm.lower() == 'dbscan':
        eps = algorithm_params.get('eps', 0.5)
        min_samples = algorithm_params.get('min_samples', 5)
        labels, clustering_params = apply_dbscan_clustering(reduced_embeddings, eps, min_samples)

    elif algorithm.lower() == 'lda':
        if k_value is None:
            raise ValueError("k_value required for LDA")
        labels, clustering_params = apply_lda_clustering(reduced_embeddings, k_value)

    elif algorithm.lower() == 'agglomerative':  # <--- AÑADIR ESTE BLOQUE
        if k_value is None:
            raise ValueError("k_value required for Agglomerative Clustering")
        labels, clustering_params = apply_agglomerative_clustering(reduced_embeddings, k_value)

    else:
        raise ValueError(f"Unknown clustering algorithm: {algorithm}")

    # Calculate metrics
    silhouette, calinski_harabasz, davies_bouldin = calculate_clustering_metrics(reduced_embeddings, labels)

    # Count actual clusters found
    unique_labels = set(labels)
    n_clusters_found = len(unique_labels) - (1 if -1 in unique_labels else 0)

    total_time = (datetime.now() - start_time).total_seconds()

    # Combine all parameters
    all_params = {
        'reduction_params': reduction_params,
        'clustering_params': clustering_params,
        'algorithm_params': algorithm_params
    }

    return ClusteringResult(
        algorithm=algorithm,
        reduction_method=reduction_method,
        n_components=n_components,
        k_value=k_value,
        cluster_labels=labels,
        reduced_embeddings=reduced_embeddings if reduction_method.lower() != 'none' else None,
        original_embeddings=embeddings,
        silhouette_score=silhouette,
        calinski_harabasz_score=calinski_harabasz,
        davies_bouldin_score=davies_bouldin,
        n_clusters_found=n_clusters_found,
        execution_time=total_time,
        parameters=all_params,
        combination_id=combination_id
    )


def save_cluster_assignments(result: ClusteringResult, df: pd.DataFrame,
                             output_path: Path, text_column: str = 'text') -> None:
    """Save cluster assignments for a specific combination."""
    clusters_dir = output_path / 'cluster_assignments'
    clusters_dir.mkdir(exist_ok=True)

    # Handle multiple membership case
    if isinstance(result.cluster_labels, list) and isinstance(result.cluster_labels[0], list):
        # Create expanded DataFrame for multiple memberships
        expanded_data = []
        for i, cluster_list in enumerate(result.cluster_labels):
            text_content = df[text_column].iloc[i]
            for cluster in cluster_list:
                expanded_data.append({
                    'original_index': i,
                    'text': text_content,
                    'cluster': cluster,
                    'combination_id': result.combination_id,
                    'is_multiple_membership': len(cluster_list) > 1
                })

        cluster_df = pd.DataFrame(expanded_data)
    else:
        # Single membership case (original behavior)
        cluster_df = pd.DataFrame({
            'original_index': range(len(df)),
            'text': df[text_column].values,
            'cluster': result.cluster_labels,
            'combination_id': result.combination_id,
            'is_multiple_membership': False
        })

    # Save to CSV
    filename = f"{result.combination_id}_clusters.csv"
    cluster_df.to_csv(clusters_dir / filename, index=False)


def create_detailed_excel_report(results: List[ClusteringResult], output_path: Path) -> None:
    """Create comprehensive Excel report with multiple sheets."""
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # 1. Summary Sheet
    ws_summary = wb.create_sheet("Summary")
    summary_data = []

    for result in results:
        summary_data.append({
            'Combination_ID': result.combination_id,
            'Algorithm': result.algorithm,
            'Reduction_Method': result.reduction_method,
            'N_Components': result.n_components,
            'K_Value': result.k_value,
            'N_Clusters_Found': result.n_clusters_found,
            'Silhouette_Score': result.silhouette_score,
            'Calinski_Harabasz_Score': result.calinski_harabasz_score,
            'Davies_Bouldin_Score': result.davies_bouldin_score,
            'Execution_Time': result.execution_time
        })

    summary_df = pd.DataFrame(summary_data)

    # Add headers and data
    for r in dataframe_to_rows(summary_df, index=False, header=True):
        ws_summary.append(r)

    # Format headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws_summary[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Auto-adjust column widths
    for column in ws_summary.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_summary.column_dimensions[column_letter].width = adjusted_width

    # 2. Best Results Sheet
    ws_best = wb.create_sheet("Best_Results")

    # Best by each metric
    best_silhouette = sorted(results, key=lambda x: x.silhouette_score, reverse=True)[:10]
    best_ch = sorted(results, key=lambda x: x.calinski_harabasz_score, reverse=True)[:10]
    best_db = sorted(results, key=lambda x: x.davies_bouldin_score)[:10]

    ws_best.append(["BEST BY SILHOUETTE SCORE"])
    ws_best.append(["Rank", "Combination_ID", "Algorithm", "Reduction", "Score"])
    for i, result in enumerate(best_silhouette, 1):
        ws_best.append([i, result.combination_id, result.algorithm,
                        result.reduction_method, result.silhouette_score])

    ws_best.append([])  # Empty row
    ws_best.append(["BEST BY CALINSKI-HARABASZ SCORE"])
    ws_best.append(["Rank", "Combination_ID", "Algorithm", "Reduction", "Score"])
    for i, result in enumerate(best_ch, 1):
        ws_best.append([i, result.combination_id, result.algorithm,
                        result.reduction_method, result.calinski_harabasz_score])

    ws_best.append([])  # Empty row
    ws_best.append(["BEST BY DAVIES-BOULDIN SCORE (Lower is Better)"])
    ws_best.append(["Rank", "Combination_ID", "Algorithm", "Reduction", "Score"])
    for i, result in enumerate(best_db, 1):
        ws_best.append([i, result.combination_id, result.algorithm,
                        result.reduction_method, result.davies_bouldin_score])

    # 3. Algorithm Performance Sheet
    ws_algo = wb.create_sheet("Algorithm_Performance")

    # Group by algorithm
    algo_performance = {}
    for result in results:
        if result.algorithm not in algo_performance:
            algo_performance[result.algorithm] = {
                'count': 0,
                'avg_silhouette': 0,
                'avg_ch': 0,
                'avg_db': 0,
                'avg_time': 0
            }
        algo_performance[result.algorithm]['count'] += 1
        algo_performance[result.algorithm]['avg_silhouette'] += result.silhouette_score
        algo_performance[result.algorithm]['avg_ch'] += result.calinski_harabasz_score
        algo_performance[result.algorithm]['avg_db'] += result.davies_bouldin_score
        algo_performance[result.algorithm]['avg_time'] += result.execution_time

    # Calculate averages
    for algo in algo_performance:
        count = algo_performance[algo]['count']
        algo_performance[algo]['avg_silhouette'] /= count
        algo_performance[algo]['avg_ch'] /= count
        algo_performance[algo]['avg_db'] /= count
        algo_performance[algo]['avg_time'] /= count

    ws_algo.append(["Algorithm", "Count", "Avg_Silhouette", "Avg_CH", "Avg_DB", "Avg_Time"])
    for algo, stats in algo_performance.items():
        ws_algo.append([algo, stats['count'], stats['avg_silhouette'],
                        stats['avg_ch'], stats['avg_db'], stats['avg_time']])

    # Save Excel file
    excel_file = output_path / 'clustering_evaluation_report.xlsx'
    wb.save(excel_file)
    logger.info(f"Excel report saved to {excel_file}")


def create_comparison_visualizations(results: List[ClusteringResult], output_path: Path) -> None:
    """Create comprehensive comparison visualizations."""
    viz_dir = output_path / 'visualizations'
    viz_dir.mkdir(exist_ok=True)

    # Prepare data for visualization
    viz_data = []
    for result in results:
        viz_data.append({
            'combination_id': result.combination_id,
            'algorithm': result.algorithm,
            'reduction_method': result.reduction_method,
            'n_components': result.n_components,
            'k_value': result.k_value,
            'n_clusters_found': result.n_clusters_found,
            'silhouette_score': result.silhouette_score,
            'calinski_harabasz_score': result.calinski_harabasz_score,
            'davies_bouldin_score': result.davies_bouldin_score,
            'execution_time': result.execution_time
        })

    df_viz = pd.DataFrame(viz_data)

    # 1. Algorithm Performance Comparison
    fig = make_subplots(
        rows=2, cols=2,
        subplot_titles=('Silhouette Score by Algorithm', 'Calinski-Harabasz Score by Algorithm',
                        'Davies-Bouldin Score by Algorithm', 'Execution Time by Algorithm'),
        specs=[[{'secondary_y': False}, {'secondary_y': False}],
               [{'secondary_y': False}, {'secondary_y': False}]]
    )

    algorithms = df_viz['algorithm'].unique()
    colors = px.colors.qualitative.Set1[:len(algorithms)]

    for i, algo in enumerate(algorithms):
        algo_data = df_viz[df_viz['algorithm'] == algo]

        fig.add_trace(go.Box(y=algo_data['silhouette_score'], name=algo,
                             marker_color=colors[i], showlegend=True), row=1, col=1)
        fig.add_trace(go.Box(y=algo_data['calinski_harabasz_score'], name=algo,
                             marker_color=colors[i], showlegend=False), row=1, col=2)
        fig.add_trace(go.Box(y=algo_data['davies_bouldin_score'], name=algo,
                             marker_color=colors[i], showlegend=False), row=2, col=1)
        fig.add_trace(go.Box(y=algo_data['execution_time'], name=algo,
                             marker_color=colors[i], showlegend=False), row=2, col=2)

    fig.update_layout(height=800, title_text="Algorithm Performance Comparison")
    fig.write_html(viz_dir / 'algorithm_comparison.html')

    # 2. Dimensionality Reduction Impact
    fig2 = px.box(df_viz, x='reduction_method', y='silhouette_score', color='algorithm',
                  title='Impact of Dimensionality Reduction on Silhouette Score')
    fig2.write_html(viz_dir / 'dimensionality_reduction_impact.html')

    # 3. K-value vs Performance (for K-dependent algorithms)
    k_dependent = df_viz[df_viz['k_value'].notna()]
    if not k_dependent.empty:
        fig3 = px.scatter(k_dependent, x='k_value', y='silhouette_score',
                          color='algorithm', size='calinski_harabasz_score',
                          title='K-value vs Silhouette Score',
                          hover_data=['davies_bouldin_score', 'execution_time'])
        fig3.write_html(viz_dir / 'k_value_performance.html')

    # 4. Performance Matrix Heatmap
    # Create a performance matrix for top combinations
    top_results = sorted(results, key=lambda x: x.silhouette_score, reverse=True)[:20]

    matrix_data = []
    for result in top_results:
        matrix_data.append([
            result.silhouette_score,
            result.calinski_harabasz_score / 1000,  # Scale down for visualization
            1 / (result.davies_bouldin_score + 0.1)  # Invert so higher is better
        ])

    combination_labels = [result.combination_id for result in top_results]

    fig4 = go.Figure(data=go.Heatmap(
        z=np.array(matrix_data).T,
        x=combination_labels,
        y=['Silhouette', 'Calinski-Harabasz (scaled)', 'Davies-Bouldin (inverted)'],
        colorscale='Viridis'
    ))

    fig4.update_layout(
        title='Performance Matrix for Top 20 Combinations',
        xaxis_title='Combination ID',
        height=600
    )
    fig4.update_xaxes(tickangle=45)
    fig4.write_html(viz_dir / 'performance_matrix.html')

    # 5. Execution Time vs Performance Trade-off
    fig5 = px.scatter(df_viz, x='execution_time', y='silhouette_score',
                      color='algorithm', size='n_clusters_found',
                      title='Execution Time vs Performance Trade-off',
                      hover_data=['combination_id', 'reduction_method'])
    fig5.write_html(viz_dir / 'time_performance_tradeoff.html')

    # 6. Static matplotlib plots for publication
    plt.style.use('seaborn-v0_8')

    # Algorithm comparison boxplot
    fig6, axes = plt.subplots(2, 2, figsize=(15, 12))

    df_viz.boxplot(column='silhouette_score', by='algorithm', ax=axes[0, 0])
    axes[0, 0].set_title('Silhouette Score by Algorithm')
    axes[0, 0].set_xlabel('Algorithm')

    df_viz.boxplot(column='calinski_harabasz_score', by='algorithm', ax=axes[0, 1])
    axes[0, 1].set_title('Calinski-Harabasz Score by Algorithm')
    axes[0, 1].set_xlabel('Algorithm')

    df_viz.boxplot(column='davies_bouldin_score', by='algorithm', ax=axes[1, 0])
    axes[1, 0].set_title('Davies-Bouldin Score by Algorithm')
    axes[1, 0].set_xlabel('Algorithm')

    df_viz.boxplot(column='execution_time', by='algorithm', ax=axes[1, 1])
    axes[1, 1].set_title('Execution Time by Algorithm')
    axes[1, 1].set_xlabel('Algorithm')

    plt.suptitle('Algorithm Performance Comparison', fontsize=16)
    plt.tight_layout()
    plt.savefig(viz_dir / 'algorithm_comparison_static.png', dpi=300, bbox_inches='tight')
    plt.close()

    logger.info(f"Visualizations saved to {viz_dir}")


def comprehensive_clustering_evaluation(csv_path: str,
                                        output_dir: str,
                                        k_min: int = 2,
                                        k_max: int = 10,
                                        dbscan_params: Optional[List[Dict]] = None,
                                        text_column: str = 'text',
                                        embeddings_column: str = 'embeddings') -> List[ClusteringResult]:
    """
    Comprehensive evaluation of clustering algorithms on text embeddings with complete output generation.

    Args:
        csv_path: Path to CSV file with embeddings
        output_dir: Directory to save results
        k_min: Minimum number of clusters to test
        k_max: Maximum number of clusters to test
        dbscan_params: List of parameter combinations for DBSCAN
        text_column: Name of column containing text data
        embeddings_column: Name of column containing embeddings

    Returns:
        List[ClusteringResult]: All clustering results
    """
    # Create output directory
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    # Load data
    logger.info(f"Loading data from {csv_path}")
    df, embeddings = load_embeddings_data(csv_path, embeddings_column, text_column)

    # Define evaluation parameters
    algorithms = ['kmeans', 'fuzzy_cmeans', 'lda', 'agglomerative']  # K-dependent algorithms
    reduction_methods = ['none', 'pca', 'umap']
    component_options = [2, 10, 50, 100]

    # DBSCAN parameters (doesn't use k)
    if dbscan_params is None:
        dbscan_params = [
            {'eps': 0.3, 'min_samples': 5},
            {'eps': 0.5, 'min_samples': 5},
            {'eps': 0.7, 'min_samples': 5},
            {'eps': 0.5, 'min_samples': 3},
            {'eps': 0.5, 'min_samples': 10}
        ]

    all_results = []
    total_combinations = 0

    # Count total combinations for progress tracking
    for algorithm in algorithms:
        for reduction_method in reduction_methods:
            if reduction_method == 'none':
                total_combinations += (k_max - k_min + 1)
            else:
                total_combinations += len(component_options) * (k_max - k_min + 1)

    # Add DBSCAN combinations
    for reduction_method in reduction_methods:
        if reduction_method == 'none':
            total_combinations += len(dbscan_params)
        else:
            total_combinations += len(component_options) * len(dbscan_params)

    logger.info(f"Starting comprehensive evaluation: {total_combinations} total combinations")

    current_combination = 0

    # Evaluate K-dependent algorithms
    for algorithm in algorithms:
        logger.info(f"Evaluating {algorithm.upper()} algorithm")

        for reduction_method in reduction_methods:
            if reduction_method == 'none':
                component_list = [None]
            else:
                component_list = component_options

            for n_components in component_list:
                # Skip invalid component combinations
                if reduction_method != 'none' and n_components >= embeddings.shape[1]:
                    continue

                for k in range(k_min, k_max + 1):
                    current_combination += 1

                    try:
                        logger.info(f"[{current_combination}/{total_combinations}] "
                                    f"{algorithm} + {reduction_method} "
                                    f"(components: {n_components}, k: {k})")

                        result = run_clustering_combination(
                            embeddings=embeddings,
                            algorithm=algorithm,
                            reduction_method=reduction_method,
                            n_components=n_components,
                            k_value=k
                        )

                        all_results.append(result)

                        # Save cluster assignments for this combination
                        save_cluster_assignments(result, df, output_path, text_column)

                        logger.info(f"Completed - Silhouette: {result.silhouette_score:.3f}, "
                                    f"CH: {result.calinski_harabasz_score:.3f}, "
                                    f"DB: {result.davies_bouldin_score:.3f}")

                    except Exception as e:
                        logger.error(f"Error in combination {current_combination}: {str(e)}")
                        continue

    # Evaluate DBSCAN
    logger.info("Evaluating DBSCAN algorithm")

    for reduction_method in reduction_methods:
        if reduction_method == 'none':
            component_list = [None]
        else:
            component_list = component_options

        for n_components in component_list:
            # Skip invalid component combinations
            if reduction_method != 'none' and n_components >= embeddings.shape[1]:
                continue

            for dbscan_param in dbscan_params:
                current_combination += 1

                try:
                    logger.info(f"[{current_combination}/{total_combinations}] "
                                f"DBSCAN + {reduction_method} "
                                f"(components: {n_components}, eps: {dbscan_param['eps']}, "
                                f"min_samples: {dbscan_param['min_samples']})")

                    result = run_clustering_combination(
                        embeddings=embeddings,
                        algorithm='dbscan',
                        reduction_method=reduction_method,
                        n_components=n_components,
                        k_value=None,
                        algorithm_params=dbscan_param
                    )

                    all_results.append(result)

                    # Save cluster assignments for this combination
                    save_cluster_assignments(result, df, output_path, text_column)

                    logger.info(f"Completed - Clusters found: {result.n_clusters_found}, "
                                f"Silhouette: {result.silhouette_score:.3f}")

                except Exception as e:
                    logger.error(f"Error in combination {current_combination}: {str(e)}")
                    continue

    # Save all results in different formats
    logger.info("Saving comprehensive results...")

    # 1. Save pickle file (original format)
    results_file = output_path / 'clustering_results.pkl'
    with open(results_file, 'wb') as f:
        pickle.dump(all_results, f)

    # 2. Save CSV summary
    save_results_summary(all_results, output_path)

    # 3. Create detailed Excel report
    create_detailed_excel_report(all_results, output_path)

    # 4. Create comprehensive visualizations
    create_comparison_visualizations(all_results, output_path)

    # 5. Generate analysis report
    create_analysis_report(all_results, output_path)

    logger.info(f"Evaluation completed! {len(all_results)} successful combinations")
    logger.info(f"Results saved to {output_path}")
    logger.info(f"- Pickle file: {results_file}")
    logger.info(f"- CSV summary: {output_path / 'clustering_results_summary.csv'}")
    logger.info(f"- Excel report: {output_path / 'clustering_evaluation_report.xlsx'}")
    logger.info(f"- Cluster assignments: {output_path / 'cluster_assignments/'}")
    logger.info(f"- Visualizations: {output_path / 'visualizations/'}")

    return all_results


def save_results_summary(results: List[ClusteringResult], output_path: Path) -> None:
    """Save a summary of results in CSV format for easy analysis."""
    summary_data = []

    for result in results:
        summary_data.append({
            'combination_id': result.combination_id,
            'algorithm': result.algorithm,
            'reduction_method': result.reduction_method,
            'n_components': result.n_components,
            'k_value': result.k_value,
            'n_clusters_found': result.n_clusters_found,
            'silhouette_score': result.silhouette_score,
            'calinski_harabasz_score': result.calinski_harabasz_score,
            'davies_bouldin_score': result.davies_bouldin_score,
            'execution_time': result.execution_time
        })

    summary_df = pd.DataFrame(summary_data)
    summary_file = output_path / 'clustering_results_summary.csv'
    summary_df.to_csv(summary_file, index=False)
    logger.info(f"Summary saved to {summary_file}")


def create_analysis_report(results: List[ClusteringResult], output_path: Path) -> None:
    """Create a comprehensive text analysis report."""
    report_file = output_path / 'analysis_report.txt'

    with open(report_file, 'w', encoding='utf-8') as f:
        f.write("COMPREHENSIVE CLUSTERING EVALUATION REPORT\n")
        f.write("=" * 80 + "\n\n")
        f.write(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Total combinations evaluated: {len(results)}\n\n")

        # Overall statistics
        f.write("OVERALL STATISTICS\n")
        f.write("-" * 50 + "\n")

        silhouette_scores = [r.silhouette_score for r in results if r.silhouette_score > -1]
        ch_scores = [r.calinski_harabasz_score for r in results if r.calinski_harabasz_score > 0]
        db_scores = [r.davies_bouldin_score for r in results if r.davies_bouldin_score < float('inf')]

        if silhouette_scores:
            f.write(f"Silhouette Score - Mean: {np.mean(silhouette_scores):.4f}, "
                    f"Std: {np.std(silhouette_scores):.4f}, "
                    f"Max: {np.max(silhouette_scores):.4f}, "
                    f"Min: {np.min(silhouette_scores):.4f}\n")

        if ch_scores:
            f.write(f"Calinski-Harabasz Score - Mean: {np.mean(ch_scores):.2f}, "
                    f"Std: {np.std(ch_scores):.2f}, "
                    f"Max: {np.max(ch_scores):.2f}, "
                    f"Min: {np.min(ch_scores):.2f}\n")

        if db_scores:
            f.write(f"Davies-Bouldin Score - Mean: {np.mean(db_scores):.4f}, "
                    f"Std: {np.std(db_scores):.4f}, "
                    f"Max: {np.max(db_scores):.4f}, "
                    f"Min: {np.min(db_scores):.4f}\n\n")

        # Algorithm performance summary
        f.write("ALGORITHM PERFORMANCE SUMMARY\n")
        f.write("-" * 50 + "\n")

        algo_stats = {}
        for result in results:
            if result.algorithm not in algo_stats:
                algo_stats[result.algorithm] = {
                    'count': 0,
                    'silhouette_scores': [],
                    'ch_scores': [],
                    'db_scores': [],
                    'execution_times': []
                }

            algo_stats[result.algorithm]['count'] += 1
            if result.silhouette_score > -1:
                algo_stats[result.algorithm]['silhouette_scores'].append(result.silhouette_score)
            if result.calinski_harabasz_score > 0:
                algo_stats[result.algorithm]['ch_scores'].append(result.calinski_harabasz_score)
            if result.davies_bouldin_score < float('inf'):
                algo_stats[result.algorithm]['db_scores'].append(result.davies_bouldin_score)
            algo_stats[result.algorithm]['execution_times'].append(result.execution_time)

        for algo, stats in algo_stats.items():
            f.write(f"\n{algo.upper()}:\n")
            f.write(f"  Combinations tested: {stats['count']}\n")
            if stats['silhouette_scores']:
                f.write(f"  Avg Silhouette: {np.mean(stats['silhouette_scores']):.4f}\n")
            if stats['ch_scores']:
                f.write(f"  Avg Calinski-Harabasz: {np.mean(stats['ch_scores']):.2f}\n")
            if stats['db_scores']:
                f.write(f"  Avg Davies-Bouldin: {np.mean(stats['db_scores']):.4f}\n")
            f.write(f"  Avg Execution Time: {np.mean(stats['execution_times']):.3f}s\n")

        # Top performing combinations
        f.write("\n\nTOP PERFORMING COMBINATIONS\n")
        f.write("-" * 50 + "\n")

        # Best by Silhouette
        best_silhouette = sorted([r for r in results if r.silhouette_score > -1],
                                 key=lambda x: x.silhouette_score, reverse=True)[:5]
        f.write("\nTop 5 by Silhouette Score:\n")
        for i, result in enumerate(best_silhouette, 1):
            f.write(f"{i}. {result.combination_id} - Score: {result.silhouette_score:.4f}\n")
            f.write(f"   Algorithm: {result.algorithm}, Reduction: {result.reduction_method}\n")
            f.write(f"   Components: {result.n_components}, K: {result.k_value}\n")

        # Best by Calinski-Harabasz
        best_ch = sorted([r for r in results if r.calinski_harabasz_score > 0],
                         key=lambda x: x.calinski_harabasz_score, reverse=True)[:5]
        f.write("\nTop 5 by Calinski-Harabasz Score:\n")
        for i, result in enumerate(best_ch, 1):
            f.write(f"{i}. {result.combination_id} - Score: {result.calinski_harabasz_score:.2f}\n")
            f.write(f"   Algorithm: {result.algorithm}, Reduction: {result.reduction_method}\n")
            f.write(f"   Components: {result.n_components}, K: {result.k_value}\n")

        # Best by Davies-Bouldin
        best_db = sorted([r for r in results if r.davies_bouldin_score < float('inf')],
                         key=lambda x: x.davies_bouldin_score)[:5]
        f.write("\nTop 5 by Davies-Bouldin Score (lower is better):\n")
        for i, result in enumerate(best_db, 1):
            f.write(f"{i}. {result.combination_id} - Score: {result.davies_bouldin_score:.4f}\n")
            f.write(f"   Algorithm: {result.algorithm}, Reduction: {result.reduction_method}\n")
            f.write(f"   Components: {result.n_components}, K: {result.k_value}\n")

        # Recommendations
        f.write("\n\nRECOMMENDATIONS\n")
        f.write("-" * 50 + "\n")

        if best_silhouette:
            best_overall = best_silhouette[0]
            f.write(f"Best overall combination based on Silhouette Score:\n")
            f.write(f"- Combination ID: {best_overall.combination_id}\n")
            f.write(f"- Algorithm: {best_overall.algorithm}\n")
            f.write(f"- Dimensionality Reduction: {best_overall.reduction_method}\n")
            f.write(f"- Number of Components: {best_overall.n_components}\n")
            f.write(f"- K Value: {best_overall.k_value}\n")
            f.write(f"- Clusters Found: {best_overall.n_clusters_found}\n")
            f.write(f"- Silhouette Score: {best_overall.silhouette_score:.4f}\n")
            f.write(f"- Execution Time: {best_overall.execution_time:.3f}s\n")

        f.write("\nFor detailed visualizations and cluster assignments, ")
        f.write("check the 'visualizations' and 'cluster_assignments' directories.\n")

    logger.info(f"Analysis report saved to {report_file}")


def analyze_best_results(results: List[ClusteringResult], top_n: int = 5) -> None:
    """Analyze and display the best clustering results."""
    if not results:
        print("No results to analyze")
        return

    print("\n" + "=" * 80)
    print("ANÁLISIS DE MEJORES RESULTADOS")
    print("=" * 80)

    # Mejores por Silhouette Score (más alto es mejor)
    valid_results = [r for r in results if r.silhouette_score > -1]
    if valid_results:
        best_silhouette = sorted(valid_results, key=lambda x: x.silhouette_score, reverse=True)[:top_n]
        print(f"\nTOP {top_n} - MEJOR SILHOUETTE SCORE:")
        print("-" * 50)
        for i, result in enumerate(best_silhouette, 1):
            print(f"{i}. {result.combination_id}")
            print(f"   {result.algorithm} + {result.reduction_method} "
                  f"(k={result.k_value}, components={result.n_components})")
            print(f"   Silhouette: {result.silhouette_score:.4f}, "
                  f"CH: {result.calinski_harabasz_score:.2f}, "
                  f"DB: {result.davies_bouldin_score:.4f}")

    # Mejores por Calinski-Harabasz (más alto es mejor)
    valid_ch_results = [r for r in results if r.calinski_harabasz_score > 0]
    if valid_ch_results:
        best_ch = sorted(valid_ch_results, key=lambda x: x.calinski_harabasz_score, reverse=True)[:top_n]
        print(f"\nTOP {top_n} - MEJOR CALINSKI-HARABASZ SCORE:")
        print("-" * 50)
        for i, result in enumerate(best_ch, 1):
            print(f"{i}. {result.combination_id}")
            print(f"   {result.algorithm} + {result.reduction_method} "
                  f"(k={result.k_value}, components={result.n_components})")
            print(f"   CH: {result.calinski_harabasz_score:.2f}, "
                  f"Silhouette: {result.silhouette_score:.4f}, "
                  f"DB: {result.davies_bouldin_score:.4f}")

    # Mejores por Davies-Bouldin (más bajo es mejor)
    valid_db_results = [r for r in results if r.davies_bouldin_score < float('inf')]
    if valid_db_results:
        best_db = sorted(valid_db_results, key=lambda x: x.davies_bouldin_score)[:top_n]
        print(f"\nTOP {top_n} - MEJOR DAVIES-BOULDIN SCORE (menor es mejor):")
        print("-" * 50)
        for i, result in enumerate(best_db, 1):
            print(f"{i}. {result.combination_id}")
            print(f"   {result.algorithm} + {result.reduction_method} "
                  f"(k={result.k_value}, components={result.n_components})")
            print(f"   DB: {result.davies_bouldin_score:.4f}, "
                  f"Silhouette: {result.silhouette_score:.4f}, "
                  f"CH: {result.calinski_harabasz_score:.2f}")


# Example usage
if __name__ == "__main__":
    # Example parameters
    csv_file = "../data/swiftkey_informative_with_embeddings (nomic).csv"
    output_directory = "../data/clustering_evaluation_results (facebook-nomic)"

    try:
        results = comprehensive_clustering_evaluation(
            csv_path=csv_file,
            output_dir=output_directory,
            k_min=10,
            k_max=30,
            text_column='Review',  # Specify your text column name
            embeddings_column='embeddings'  # Specify your embeddings column name
        )

        print(f"Evaluation completed with {len(results)} combinations")
        print(f"Results saved to: {output_directory}")
        print("\nFiles generated:")
        print("- clustering_results.pkl (complete results)")
        print("- clustering_results_summary.csv (metrics summary)")
        print("- clustering_evaluation_report.xlsx (detailed Excel report)")
        print("- analysis_report.txt (comprehensive text analysis)")
        print("- cluster_assignments/ (individual cluster assignments)")
        print("- visualizations/ (comparison charts and plots)")

        # Show best results by metric
        analyze_best_results(results, top_n=3)

    except Exception as e:
        print(f"Error: {e}")
        import traceback

        traceback.print_exc()
